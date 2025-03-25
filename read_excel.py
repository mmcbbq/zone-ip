import openpyxl

path = 'Mappe1.xlsx'

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
ip = 10
teil = []
file = open('name_ip.txt', 'w')
for j in range(2, max_row + 1):
    sub = sheet_obj.cell(row=j, column=1).value[0:2] + sheet_obj.cell(row=j, column=2).value[0]
    c = 1
    while sub in teil:
        sub = sheet_obj.cell(row=j, column=1).value[0 + 1:2 + 1] + sheet_obj.cell(row=j, column=2).value[0]
        c += 1
    teil.append(sub.lower())

    file.write(f'Name: {sheet_obj.cell(row=j, column=1).value} {sheet_obj.cell(row=j, column=2).value}\n')
    file.write(f"Domainname: {sub.lower()}.linux.zz\n")
    file.write(f'Netzwerk: eth0 192.168.{ip}.0/24\n')
    file.write(f'Netzwerk: eth1 10.101.214.{ip + 100}/24\n')
    file.write(f'Router ip: eth0 192.168.{ip}.254/24\n')
    file.write(f'server-01 ip: eth0 192.168.{ip}.250/24\n')
    file.write(f'server-02 ip: eth0 192.168.{ip}.251/24\n')
    file.write(f'client ip: eth0 192.168.{ip}.1/24\n\n\n\n')

    ip += 5
file.close()
print(teil)
