import openpyxl

path = 'Mappe1.xlsx'

wb_obj = openpyxl.load_workbook(path)
sheet = wb_obj.active

# sheet_obj = wb_obj.active

max_col = sheet.max_column
max_row = sheet.max_row
ip = 10

for x in range(2,max_row):
    print(f'        - to {sheet.cell(row=x, column=4).value[0:-6]}0/24')
    print(f'          via {sheet.cell(row=x, column=5).value[:-3]}')
    print()