import openpyxl
# import subprocess
from ping3 import ping
# host = '10.101.214.254'
#
# time = ping(host)



path = 'Mappe1.xlsx'

wb_obj = openpyxl.load_workbook(path)
sheet = wb_obj.active

# sheet_obj = wb_obj.active

max_col = sheet.max_column
max_row = sheet.max_row
ip = 10
teil = []
file = open('name_ip.txt', 'w')
for j in range(2, max_row + 1):
    host = sheet.cell(row=j,column=4).value
    time = ping(host)

    print(sheet.cell(row=j,column=2).value)
    print(sheet.cell(row=j,column=4).value)
    print(time)
    print()